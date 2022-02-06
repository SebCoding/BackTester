import datetime as dt
from datetime import timedelta

import dask.dataframe as dd
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

import constants

# Adjust from_time to include prior X entries for that interval for ema200
from Configuration import Configuration


def adjust_from_time(from_time, interval, include_prior):
    if interval not in constants.VALID_INTERVALS:
        raise Exception(f'Invalid interval value: {interval}')

    delta = include_prior - 1
    if 'm' in interval:
        interval = interval.replace('m', '')
        from_time = from_time - timedelta(minutes=int(interval) * delta)
    elif 'h' in interval:
        interval = interval.replace('h', '')
        from_time = from_time - timedelta(hours=int(interval) * delta)
    elif 'd' in interval:
        from_time = from_time - timedelta(days=delta)
    elif 'w' in interval:
        from_time = from_time - timedelta(weeks=delta)
    return from_time


# Convert an index value of type numpy.datetime64 to type datetime
def idx2datetime(index_value):
    return dt.datetime.utcfromtimestamp(index_value.astype('O') / 1e9)


def save_trades_to_file(test_num, exchange, pair, from_time, to_time, interval, df, include_time=False, verbose=True):
    config = Configuration.get_config()
    test_num = str(test_num)
    pair = pair.replace('/', '-')

    if include_time:
        from_str = from_time.strftime('%Y-%m-%d %H.%M')
        to_str = to_time.strftime('%Y-%m-%d %H.%M')
    else:
        from_str = from_time.strftime('%Y-%m-%d')
        to_str = to_time.strftime('%Y-%m-%d')

    filename = f'{exchange} {pair} [{interval}] {from_str} to {to_str}'
    filename = f"{config['output']['results_path']}\\{test_num} {filename} Trades"

    if 'csv' in config['output']['output_file_format']:
        filename = filename + '.csv'
        df.to_csv(filename, index=True, header=True)
        if verbose:
            print(f'Trades file created => [{filename}]')
    if 'xlsx' in config['output']['output_file_format']:
        filename = filename + '.xlsx'
        df.to_excel(filename, index=True, header=True)
        # to_excel_formatted(df, filename)
        if verbose:
            print(f'Trades file created => [{filename}]')


# TODO: Find a way to format the Excel workbook prior to saving to file
def to_excel_formatted(df, filename):
    wb = Workbook()
    ws = wb.active

    for r in dataframe_to_rows(df, index=True, header=True):
        ws.append(r)

    # TODO: Fix this code. It destroys index datetime format
    for cell in ws['A'] + ws[1]:
        cell.style = 'Pandas'

    wb.save(filename)


def convert_interval_to_min(interval):
    if interval not in constants.VALID_INTERVALS:
        raise Exception(f'Invalid interval value: {interval}')

    if 'm' in interval:
        interval = interval.replace('m', '')
        return int(interval)
    elif 'h' in interval:
        interval = interval.replace('h', '')
        return int(interval) * 60
    elif 'd' in interval:
        interval = interval.replace('d', '')
        return int(interval) * 1440
    elif 'w' in interval:
        interval = interval.replace('w', '')
        return int(interval) * 10080
    else:
        return 0


def read_excel_to_dataframe(filename):
    wb = load_workbook(filename)
    ws = wb['Sheet1']

    # To convert a worksheet to a Dataframe you can use the value's property.
    # This is very easy if the worksheet has no headers or indices:
    # df = DataFrame(ws.values)

    # https://openpyxl.readthedocs.io/en/stable/pandas.html

    # If the worksheet does have headers or indices, such as one created by Pandas,
    # then a little more work is required:
    from itertools import islice
    data = ws.values
    cols = next(data)[1:]
    data = list(data)
    idx = [r[0] for r in data]
    data = (islice(r, 1, None) for r in data)
    df = pd.DataFrame(data, index=idx, columns=cols)

    return df


def read_csv_to_dataframe(filename):
    dask_df = dd.read_csv(filename, parse_dates=['Unnamed: 0']).set_index('Unnamed: 0')
    df = dask_df.compute()
    df.index.name = None
    # print(f'from:{from_time} to:{to_time}')
    # print(df.to_string())
    return df


def read_csv_to_dataframe_by_range(filename, from_time, to_time):
    dask_df = dd.read_csv(filename, parse_dates=['Unnamed: 0']).set_index('Unnamed: 0')
    dask_df = dask_df.loc[from_time:to_time]
    df = dask_df.compute()
    df.dropna(inplace=True)
    df.index.name = None
    # print(f'from:{from_time} to:{to_time}')
    # print(df.to_string())
    return df


def format_execution_time(seconds):
    # Remove days and keep remainder in seconds
    seconds = seconds % (24 * 3600)
    hours = seconds // 3600
    seconds %= 3600
    minutes = seconds // 60
    seconds %= 60
    output = f"{int(hours)}h {int(minutes)}m {int(seconds)}s"
    for c in output:
        if c in ['0', ':', 'h', 'm', 's', ' ']:
            output = output.replace(c, '', 1)
        else:
            break
    if len(output) == 0:
        output = 'less than 1s'
    return output


def isNan(value):
    """
        The usual way to test for a NaN is to see if it's equal to itself:
        https://stackoverflow.com/questions/944700/how-can-i-check-for-nan-values
    """
    return value != value
